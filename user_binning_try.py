# -*- coding: utf-8 -*-
"""
Created on Tue Apr  3 19:16:20 2018

@author: ishwarya.sriraman
"""

import pandas as pd
from eda2 import eda
import numpy as np
from openpyxl import load_workbook
import openpyxl
class user_bin:
    def __init__(self,dataframe,target,event):
        self.dataframe=dataframe
        self.target=target
        self.event=event
        
    def user_binned(self):
        try:
            fin_data=pd.DataFrame()
            df=pd.read_excel('D:\\Other projects\\python modules\\univariates_bivariates.xlsx','finecuts')
            df=df[df.columns[~df.columns.str.contains('Unnamed:')]]
#            del df['CAT']
#            del df['GOODS']
#            del df['BADS']
#            del df['TOTAL']
#            del df['PCT_G']
#            del df['PCT_B']
#            del df['WOE']
#            del df['IV']
#            del df['TOTAL_PERC']
            finecuts_cols = [col for col in df.columns if 'FineBinCuts' in col or 'user_inputs' in col]
            df1=df[finecuts_cols].dropna(axis=1, how='all')
            l=df1.columns
            print(l)
            for i in range(len(l)):
                if 'user_inputs'  in l[i]:
                    pass
                elif i==len(l)-1 and 'user_inputs' not in l[i]:
                    del df1[l[i]]
                elif 'user_inputs' not in l[i+1]:
                    del df1[l[i]]
                i=i+1
            df6=pd.DataFrame()
            print('df1 is')
            print(df1)
            for i in range(0, df1.shape[1], 2):
                    uiname = df1.columns[i+1]
                    if('user_inputs' in uiname):
                        userinp = df1[uiname].values.tolist()
                    start = userinp[0]
                    l1 =[]
                    binname = df1.columns[i]
                    if('FineBinCuts' in binname):
                        for j,val in enumerate(userinp):
                            if val!=start:
                                l1.append((df1[binname][j-1]))
                                start = val
                    df5=self.cuts_func(self.dataframe[df1.columns[i].split('_')[0]], l1)
                    print('df5 is')
                    print(df5)
                    df6[df5.name]=df5.astype('str')
                    df6[self.target]=self.dataframe[self.target]
                    print('df6 is')
                    print(df6)
                    e=eda(df6,self.target,self.event)
                    d=e.bininfo(0)
    #                mid_data=pd.DataFrame(mid_data.items())
                    a = pd.DataFrame()
                    for name in d.keys():
                        b = d[name]
                        b['keys'] = [name]*b.shape[0]
                        a = pd.concat([a, b])
    #                a=a.set_index('keys')
                    fin_data=pd.concat([fin_data,a],axis=0)
                    
            uni1=pd.read_excel('D:\\Other projects\\python modules\\univariates_bivariates.xlsx','bivariates')
            book = load_workbook('univariates_bivariates.xlsx')
            writer = pd.ExcelWriter('univariates_bivariates.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            uni1 = uni1[~uni1['keys'].isnull()]
            for i in fin_data['keys'].unique():
                uni1=uni1.drop(uni1[uni1['keys']==i].index)
            print(uni1['keys'].unique())
            k=0
            new_cols=[]
            varnames = uni1['keys'].unique()
            varnames =varnames.tolist()
            try:
                varnames.remove('keys')
            except:
                pass
            df7=pd.DataFrame()
            
            workbook=openpyxl.load_workbook('univariates_bivariates.xlsx')
            std=workbook.get_sheet_by_name('bivariates')
            workbook.remove_sheet(std)
            workbook.save('univariates_bivariates.xlsx')
            book = load_workbook('univariates_bivariates.xlsx')
#            workbook=openpyxl.load_workbook('univariates_bivariates.xlsx')
            writer = pd.ExcelWriter('univariates_bivariates.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            
            for i in varnames:
                print(i)
                try:
                    df7 = uni1[uni1['keys']==i]
               
                    print(df7)
                    if min(df7.index.tolist()) != 0:
                        df7.columns = uni1.loc[df7.index[0]-1].tolist()
                        new_cols = [x for x in df7.columns  if x is not np.nan]
                        new_df = df7[new_cols]
                        new_df.to_excel(writer,sheet_name='bivariates',startrow=k,startcol=0,index=False)
                        k= k + df7.shape[0] + 2
                        print(k)
                        df7.set_index('keys',inplace=True)
                    else:
                        df7.columns=uni1.columns
                        df7=df7[df7.columns[~df7.columns.str.contains('Unnamed:')]]
                        new_cols = [x for x in df7.columns  if x is not np.nan]
                        df7 = df7[new_cols]
                        df7.to_excel(writer,sheet_name='bivariates',startrow=k,startcol=0,index=False)
                        k= k + df7.shape[0] + 2
                        df7.set_index('keys',inplace=True)
                except:
                    pass
            fin_data.set_index('keys',inplace=True)   
            print(fin_data)
            fin_data.to_excel(writer,sheet_name='bivariates',startrow=k,startcol=0,index=True)
            writer.save()
            writer.close()
        except Exception as e:
            print(e)
            

    def cuts_func(self,vals, cuts):
            leftend = vals.min()-1
            rightend = vals.max()
            edgevalues = cuts
            cut_vals = [leftend]+edgevalues+[rightend]
            res = pd.cut(vals, cut_vals,labels=["<=" +str(x) for x in cut_vals[1:]])
            return(res)
