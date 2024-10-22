# -*- coding: utf-8 -*-
"""
Created on Tue Mar 20 12:22:27 2018

@author: ishwarya.sriraman
"""
from rpy2.robjects import pandas2ri
from rpy2.robjects.packages import STAP
import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
from eda2 import eda
class woe:
    def __init__(self,df,target,event):
        self.df=df
        self.target=target
        self.event=event

    def fine_cuts(self):
        while True:
            try:
                with open('woe_final.r', 'r') as f:
                    string = f.read()
                    woe_final = STAP(string, "woe_final")
            except FileNotFoundError:
                path = input("set directory to path: ")
                from os import chdir
                chdir(path)
                with open('woe_final.r', 'r') as f:
                    string = f.read()
                    woe_final = STAP(string, "woe_final")
            else:
                break
            
        pandas2ri.activate()
        data=woe_final.finebincuts(self.df,self.target,self.event)
        pandas2ri.deactivate()
        pandas2ri.activate()
        data1=woe_final.woebincuts(self.df,self.target,self.event)
        pandas2ri.deactivate()
        try:
            book = load_workbook('univariates_bivariates.xlsx')
            writer = pd.ExcelWriter('univariates_bivariates.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            k=0
            val_list=[]
            range_list=[]
            for i in range(len(data)):
                b=pandas2ri.ri2py_dataframe(data[i])
                b.columns=[data.names[i]+'_FineBinCuts']
                val_list.append(b)
                b1=pandas2ri.ri2py_dataframe(data1[i])
                b1['name']=data1.names[i]
                range_list.append(b)
                b['assigned_cuts']=self.cuts_func(b[data.names[i]+'_FineBinCuts'], b1.WOECuts)
#                b.set_index('name',inplace=True)  
                l=[]
                l=list(b[data.names[i]+'_FineBinCuts'])
                l.append((self.df[data.names[i]].min())-1)
                l.append(self.df[data.names[i]].max())
                l=list(set(l))
                l=sorted(l)
                dat=pd.cut(self.df[data.names[i]],l,labels=["<=" +str(x) for x in l[1:]])
                col=data.names[i]+'_FineBinCuts'
                df1=pd.DataFrame(data=dat)
                df1.columns=[col]
                df1[self.target]=self.df[self.target]
                df1[col]=df1[col].astype('object')
#                print('df1 is')
#                print(df1.head())
#                print('df1 type')
#                print(type(df1[col].dtypes))
                e=eda(df1,self.target,self.event)
                d=e.bininfo(0)
#                print('d is')
#                print(d)
                a=pd.DataFrame()
                for name in d.keys():
                        a = d[name]
                        a['keys'] = [name]*a.shape[0]
                b=pd.concat([a,b],axis=1)
                del b['keys']
#                del b[col]
#                b.rename(columns={'CAT':col})
#                b=b[['age_FineBinCuts','assigned_cuts','CAT','GOODS','BADS','TOTAL','	PCT_G','PCT_B','WOE','IV','TOTAL_PERC']]
                b['CAT']=b['CAT'].apply(lambda x:int(x[2:]))
                b['CAT']=sorted(b['CAT'])
                b['CAT']=b['CAT'].apply(lambda x:'<='+str(x))
                st=b['CAT'].iloc[b.shape[0]-1]
                p=st.split('<=')[1]
                b[data.names[i]+'_FineBinCuts'].iloc[b.shape[0]-1]=int(p)
                b['user_inputs']=np.nan
                b['assigned_cuts'].iloc[len(b['assigned_cuts'])-1]=b['assigned_cuts'].iloc[len(b['assigned_cuts'])-2]
                b.to_excel(writer,sheet_name='finecuts',startrow=1,startcol=k,index=False)
                k=k+b.shape[1]+2
            writer.close()
        except:
            book = load_workbook('univariates_bivariates.xlsx')
            writer = pd.ExcelWriter('univariates_bivariates.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            k=0
            val_list=[]
            range_list=[]
            for i in range(len(data)):
                b=pandas2ri.ri2py_dataframe(data[i])
                b.columns=[data.names[i]+'_FineBinCuts']
                print(data.names[i])
                print(b.columns)
                b1=pandas2ri.ri2py_dataframe(data1[i])
                b1['name']=data1.names[i]
                b['assigned_cuts']=self.cuts_func(b[data.names[i]+'_FineBinCuts'], b1.WOECuts)
                print(b['assigned_cuts'].dtypes)
                l=[]
                l=list(self.df[data.names[i]])
                l.append(self.df[data.names[i]].min()-1)
                l.append(self.df[data.names[i]].max())
                l=list(set(l))
                l=sorted(l)
                dat=pd.cut(self.df[data.names[i]],l,labels=["<=" +str(x) for x in l[1:]])
                col=data.names[i]+'_FineBinCuts'
                df1=pd.DataFrame(data=dat)
                df1.columns=[col]
                df1[self.target]=self.df[self.target]
                df1[col]=df1[col].astype('object')
#                print('df1 is')
#                print(df1.head())
#                print('df1 type')
#                print(type(df1[col].dtypes))
                e=eda(df1,self.target,self.event)
                d=e.bininfo(0)
#                print('d is')
#                print(d)
                a=pd.DataFrame()
                for name in d.keys():
                        a = d[name]
                        a['keys'] = [name]*a.shape[0]
                b=pd.concat([a,b],axis=1)
                try:
                    del b['keys']
                except:
                    print("")
#                b.set_index('name',inplace=True)
#                del b[col]
#                b.rename(columns={'CAT':col})
                #b=b[['age_FineBinCuts','assigned_cuts','CAT','GOODS','BADS','TOTAL','	PCT_G','PCT_B','WOE','IV','TOTAL_PERC']]
                b['user_inputs']=np.nan
                b['CAT']=b['CAT'].apply(lambda x:int(x[2:]))
                b['CAT']=sorted(b['CAT'])
                b['CAT']=b['CAT'].apply(lambda x:'<='+str(x))
                st=b['CAT'].iloc[b.shape[0]-1]
                p=st.split('<=')[1]
                b[data.names[i]+'_FineBinCuts'].iloc[b.shape[0]-1]=int(p)
                b['assigned_cuts'].iloc[len(b['assigned_cuts'])-1]=b['assigned_cuts'].iloc[len(b['assigned_cuts'])-2]
                b.to_excel(writer,sheet_name='finecuts',startrow=1,startcol=k,index=False)    
                k=k+b.shape[1]+2
            writer.close()

    def cuts_func(self,vals, cuts):
        leftend = vals.min()-1
        rightend = vals.max()
        edgevalues = cuts.tolist()
        cut_vals = [leftend]+edgevalues+[rightend]
        print(cut_vals)
        res = pd.cut(vals, cut_vals, labels=range(len(cut_vals)-1))
        return(res)

    def entropy_cuts(self):
        while True:
            try:
                with open('woe_final.r', 'r') as f:
                    string = f.read()
                    woe_final = STAP(string, "woe_final")
            except FileNotFoundError:
                path = input("set directory to path: ")
                from os import chdir
                chdir(path)
                with open('woe_final.r', 'r') as f:
                    string = f.read()
                    woe_final = STAP(string, "woe_final")
            else:
                break
            
        pandas2ri.activate()
        data=woe_final.entropybincuts(self.df,self.target)
        pandas2ri.deactivate()
        
        try:
            os.remove('D:/Other projects/python modules/entropybincuts.xlsx')
            engine = 'xlsxwriter'
            writer = pd.ExcelWriter('entropybincuts.xlsx',engine=engine)
            k=0
            for i in range(len(data)):
                b=pandas2ri.ri2py_dataframe(data[i])
                b['name']=data.names[i]
                b.set_index('name',inplace=True)
                b.to_excel(writer,sheet_name='entropybincuts',startrow=1,startcol=k,index=True)
                k=k+b.shape[1]+2
            writer.close()
        except:
            engine = 'xlsxwriter'
            writer = pd.ExcelWriter('entropybincuts.xlsx',engine=engine)
            k=0
            for i in range(len(data)):
                b=pandas2ri.ri2py_dataframe(data[i])
                b['name']=data.names[i]
                b.set_index('name',inplace=True)
                b.to_excel(writer,sheet_name='entropybincuts',startrow=1,startcol=k,index=True)
                k=k+b.shape[1]+2
            writer.close()
        a=pd.DataFrame()
        for i in range(len(data)):
            b=pandas2ri.ri2py_dataframe(data[i])
            b['name']=data.names[i]
            a=pd.concat([a,b],axis=0)
        a.set_index('name',inplace=True)
        return a