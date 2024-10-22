import numpy as np

import pandas as pd

from sklearn.preprocessing import LabelEncoder

from scipy.stats import kurtosis

from scipy.stats import skew

import scipy as sc

import missingno as msno

from rpy2.robjects import r, pandas2ri

from rpy2.robjects.packages import STAP

import rpy2.robjects as robjects

from rpy2.robjects.vectors import StrVector

import rpy2.robjects as ro

from rpy2.robjects.packages import importr

import rpy2.robjects.packages as rpackages

from rpy2.robjects.lib.dplyr import dplyr

from rpy2.robjects import default_converter

from rpy2.robjects.conversion import localconverter

import matplotlib as mpl

import matplotlib.pyplot as plt

import seaborn as sns

import re

from sklearn.feature_selection import VarianceThreshold

import os

from openpyxl import load_workbook

from statsmodels.stats.outliers_influence import variance_inflation_factor

import math

 

class eda:

    ''' This is the module designed for exploratory data analysis. Using this module User

    can do univariate, bi-variate analysis. Along with it user may get binning sugestion, can

    remove variable univariate way or bivariate way'''

    def __init__(self,data,target,event):

       

        ''' This method is initialize class eda object'''

#        self.pool =  mp.Pool(processes=2)

        self.train = data

        self.target = data[target]

        self.targetVarName = target

        self.event = event

   

    def univariate(self):

        '''

            this function is for univariate analysis,

        '''

        #if 

        x = [x for x in self.train.columns if self.train[x].dtypes != 'object']

        data = self.train[x]

        #y = data.describe()

        evcount = self.target.value_counts()

        evratio = evcount[self.event]/self.target.count()

        def Range(y):

            mi = np.min(y)

            ma = np.max(y)

            rng = ma-mi

            return rng

        Rnge = data.apply(Range)

        def iqr(y):

            x25 = np.percentile(y,25)

            x75 = np.percentile(y,75)

            Iqr = x75 - x25

            return Iqr

           

        iQr = data.apply(iqr)

        def Skew(y):

            d = skew(y)

            return d

        skewness = data.apply(Skew)

        def Kurt(y):

            d = kurtosis(y)

            return d

        Kurtosis = data.apply(Kurt)

       

        mean = data.apply(np.mean)

        stdv = data.apply(np.std)

        median = data.apply(np.median)

        missingno = data.isnull().sum()

        def outlier(y):

            q1 = np.percentile(y,25)

            q3 = np.percentile(y,75)

            iqr = q3-q1

            outlier = ((y > (q3 +1.5*iqr)) |(y <(q1-1.5*iqr ))).sum()

            return outlier

        Q1 = data.quantile(.25) 
        
        p1=data.quantile(0.01)
        p3=data.quantile(0.99)

        Q3 = data.quantile(.75)

        out = data.apply(outlier)
         
        count=self.train.shape[0]
        
        missingper=(missingno/count)*100

        Frame = [Rnge,Q1,Q3,p1,p3,iQr,skewness,Kurtosis,mean,stdv,median,missingno,out,missingper]

        unistat = pd.concat(Frame,axis =1)

        unistat.columns = ['Range','Q1','Q3','p1','p99','IQR','Skewness','Kurtosis','mean','stdv','median','MissingValue','outlier','Missing_perc']

        l=[]
        
        for i in missingper:
            if i>99:
                l.append(1)
            else:
                l.append(0)
                
        unistat['Missing_Criteria']=l
        
        unistat=unistat[['MissingValue','Missing_perc','mean','median','stdv','p1','p99','Range','Q1','Q3','IQR','outlier','Skewness','Kurtosis']]
        z = [x for x in self.train.columns if self.train[x].dtypes == 'object']

        dat = self.train[z]

        cat = {}

        for i in z:

            cat[i]= dat[i].value_counts()
        
        data={'catdist':cat,'event Ratio':evratio , 'unistat':unistat}

#        print('data is')
#        print(data)

        try:

            k=0

            os.remove('D:/Other projects/python modules/univariates_bivariates.xlsx')

            engine = 'xlsxwriter'

            writer = pd.ExcelWriter('univariates_bivariates.xlsx',engine=engine)

            uni=data['unistat']

            k=len(uni)+1

            uni.to_excel(writer,sheet_name='univariates',startrow=0,startcol=0,index=True)
            
#            print(data)
            for i in data['catdist']:
              sum1=self.train.shape[0]
              l1=[]
              report_df = pd.DataFrame()
              print(data['catdist'][i].index)
              report_df[i] = list(data['catdist'][i].index)
              report_df["Numbers"] = data['catdist'][i].values
              for j in data['catdist'][i]:
                  l1.append(j/sum1)
              report_df["perc"] = l1
              #print('data of catdist is')
              #print(data["catdist"][i])
#              data['catdist'][i]['perc']=l1
              report_df.to_excel(writer,sheet_name='univariates',startrow=k+1,startcol=0,index=True)

              k=len(data['catdist'][i])+2+k

            writer.close()

        except:

            k=0

            engine = 'xlsxwriter'

            writer = pd.ExcelWriter('univariates_bivariates.xlsx',engine=engine)

            uni=data['unistat']

            k=len(uni)+1

            uni.to_excel(writer,sheet_name='univariates',startrow=0,startcol=0,index=True)

            for i in data['catdist']:
              sum1=self.train.shape[0]
              report_df = pd.DataFrame()
              print(data['catdist'][i].index)
              report_df[i] = list(data['catdist'][i].index)
              report_df["Numbers"] = data['catdist'][i].values
              l1=[]
              for j in data['catdist'][i]:
                  l1.append(j/sum1)
              report_df["perc"] = l1
#              print('reprt_df is')
#              print(report_df)
              report_df.to_excel(writer,sheet_name='univariates',startrow=k+1,startcol=0,index=True)

              k=len(data['catdist'][i])+2+k

            writer.close()

        

        return {'catdist':cat,'event Ratio':evratio , 'unistat':unistat.transpose()}

   

    def correlation(self,X):

         cor=[]

         cor1=[]

         for i in X:

             cor.append(self.df[i].corr(self.df[self.target]))

         cor1 = [0 if math.isnan(x) else x for x in cor]

            

         sum1=sum(cor1)

         len1=len(cor1)

         mean=sum1/len1

         dictionary = dict(zip(cor1, X))

         print(dictionary)

         for k in list(dictionary):

             if mean>k:

                 del dictionary[k]

               

         return list(dictionary.values())

    

    

    

    def boxgraph(self):

 

        data=self.train

 

        y = self.targetVarName

 

        x = [x for x in data.columns if data[x].dtypes != 'object']

 

        img ={}

 

        for col in x:

 

            z=sns.boxplot(data[col],data[y])

 

            fig = z.get_figure()

 

            img[col]=fig

 

        return img

 

    def dist(self):

 

        data=self.train

 

        y = self.targetVarName

 

        x = [x for x in data.columns if data[x].dtypes != 'object']

 

        img ={}

 

        for col in x:

 

            z=sns.pairplot(data[[col,y]],hue=y)

 

            fig = z.fig

 

            img[col]=fig

 

        return img

   

    

    

    def metrics(self,method='IV'):

        ''' this is the function to calcualte, NMI, ChiSq stat and IV value for each algorithm.

        Continous variable are first converted in to categorical variable afgter binning process.

        Example to use:

            h= eda(data,'y','yes')

            f = h.metrics()

            it requires one argument method which is by defualt set to IV

            method arguments accepts value such as None, entropy,Chisq and IV'''

               

        utils = importr('utils')

        if method is None or method == 'entropy':

            cat = [x for x in self.train.columns if self.train[x].dtypes == 'object']

            num = [x for x in self.train.columns if self.train[x].dtypes != 'object']

            numdata = self.train[num]

            catdata = self.train[cat]

            while True:

                try:

                    smbinning = importr('smbinning')

                    woeBinning = importr('woeBinning')

                    InformationValue = importr('InformationValue')

                    foreach = importr('foreach')

                    doParallel = importr('doParallel')

                except Exception as e:

                   

                    y = str(e)

                    print(y)

                    z = re.findall(r"'(.*?)'", y, re.DOTALL)[0]

                    utils.install_packages(z)

                else:

                    break

            while True:

                try:

                    with open('woe.r', 'r') as f:

                        string = f.read()

                        woe = STAP(string, "woe")

                        print('woe module loaded')

                except FileNotFoundError:

                    path = input("set directory to path: ")

                    from os import chdir

                    chdir(path)

                    with open('woe.r', 'r') as f:

                        string = f.read()

                        woe = STAP(string, "woe")

                        print('woe module loaded')

                else:

                    break

            self.MutualI = {}

            for cal in catdata.columns:

                if cal != self.targetVarName:

                    pandas2ri.activate()

                    nmi = pandas2ri.ri2py(woe.nmi(catdata[cal],catdata[self.targetVarName]))

                    pandas2ri.deactivate()

                    self.MutualI[cal]=nmi

            numdat = numdata.copy()

            numdat[self.targetVarName]= self.target

            pandas2ri.activate()

            entdata = pandas2ri.ri2py(woe.entropy_based_bin(numdat))

            pandas2ri.deactivate()

            for cal in numdata.columns:

                if cal != self.targetVarName:

                    pandas2ri.activate()

                    nmi = pandas2ri.ri2py(woe.nmi(entdata[cal],entdata[self.targetVarName]))

                    pandas2ri.deactivate()

                    self.MutualI[cal]=nmi

                   

            return self.MutualI

        elif method == 'Chisq':

            cat = [x for x in self.train.columns if self.train[x].dtypes == 'object']

            num = [x for x in self.train.columns if self.train[x].dtypes != 'object']

            numdata = self.train[num]

            catdata = self.train[cat]

           

            while True:

                try:

                    smbinning = importr('smbinning')

                    woeBinning = importr('woeBinning')

                    InformationValue = importr('InformationValue')

                    foreach = importr('foreach')

                    doParallel = importr('doParallel')

                except Exception as e:

                   

                    y = str(e)

                    z = re.findall(r"'(.*?)'", y, re.DOTALL)[0]

                    utils.install_packages(z)

                else:

                    break

            while True:

                try:

                    with open('woe.r', 'r') as f:

                        string = f.read()

                        woe = STAP(string, "woe")

                        print('woe module loaded')

                except FileNotFoundError:

                    path = input("set directory to path: ")

                    from os import chdir

                    chdir(path)

                    with open('woe.r', 'r') as f:

                        string = f.read()

                        woe = STAP(string, "woe")

                        print('woe module loaded')

                else:

                    break

            self.Chistat = {}

            for cal in catdata.columns:

                if cal != self.targetVarName:

                    pandas2ri.activate()

                    nmi = pandas2ri.ri2py(woe.ChiTest(catdata[cal],catdata[self.targetVarName]))

                    pandas2ri.deactivate()

                    self.Chistat[cal]=nmi

            numdat = numdata.copy()

            numdat[self.targetVarName]= self.target

            pandas2ri.activate()

            entdata = pandas2ri.ri2py(woe.entropy_based_bin(numdat))

            pandas2ri.deactivate()

            for cal in numdata.columns:

                if cal != self.targetVarName:

                    pandas2ri.activate()

                    nmi = pandas2ri.ri2py(woe.ChiTest(entdata[cal],entdata[self.targetVarName]))

                    pandas2ri.deactivate()

                    self.Chistat[cal]=nmi

            return self.Chistat

        elif method == "IV":

            cat = [x for x in self.train.columns if self.train[x].dtypes == 'object']

            num = [x for x in self.train.columns if self.train[x].dtypes != 'object']

            numdata = self.train[num]

            catdata = self.train[cat]

            while True:

                try:

                    smbinning = importr('smbinning')

                    woeBinning = importr('woeBinning')

                    InformationValue = importr('InformationValue')

                    foreach = importr('foreach')

                    doParallel = importr('doParallel')

                except Exception as e:

                   

                    y = str(e)

                    print(y)

                    z = re.findall(r"'(.*?)'", y, re.DOTALL)[0]

                    utils.install_packages(z)

                else:

                    break

            while True:

                try:

                    with open('woe.r', 'r') as f:

                        string = f.read()

                        woe = STAP(string, "woe")

                        print('woe module loaded')

                except FileNotFoundError:

                    path = input("set directory to path: ")

                    from os import chdir

                    chdir(path)

                    with open('woe.r', 'r') as f:

                        string = f.read()

                        woe = STAP(string, "woe")

                        print('woe module loaded')

                else:

                    break

            self.IVtable = {}

            for cal in catdata.columns:

                if cal != self.targetVarName:

                    pandas2ri.activate()

                    nmi = pandas2ri.ri2py(woe.wtable(catdata[cal],catdata[self.targetVarName],self.event))

                    pandas2ri.deactivate()

                    self.IVtable[cal]=nmi

            numdat = numdata.copy()

            numdat[self.targetVarName]= self.target

            pandas2ri.activate()

            while True:

                #i=0

                try:

                    #i=0

                    entdata = pandas2ri.ri2py(woe.woebin(numdat,self.targetVarName,self.event))

                    print('trying monobin')

                except Exception as e:

                    y = str(e)

                    z = re.findall(r"'(.*?)'", y, re.DOTALL)[0]

                    utils.install_packages(z)

                else:

                    break

                   

            entdata[self.targetVarName]=  self.target

            pandas2ri.deactivate()

            for cal in entdata.columns:

                if cal != self.targetVarName:

                    pandas2ri.activate()

                    nmi = pandas2ri.ri2py(woe.wtable(entdata[cal],entdata[self.targetVarName],self.event))

                    pandas2ri.deactivate()

                    self.IVtable[cal]=nmi

            return self.IVtable

        else:

            print('Please choose method argument equal to one of the list of IV,Chisq, None or entropy')

           

        

         

    def varianceThreshold(self,threshold,X):

        X=self.train[X]

        datecol=[x for x in self.train.columns if self.train[x].dtypes=='datetime64[ns]']

        X=[x for x in X.columns if X[x].dtypes != 'object' and x not in datecol and x not in self.targetVarName]

        

        data=self.train[X]

        columns = data.columns

        selector = VarianceThreshold(threshold)

        s=selector.fit(data)

        s.transform(data)

        labels = [columns[x] for x in selector.get_support(indices=True)]

        vts=pd.DataFrame(selector.fit_transform(data), columns=labels)

       

        return(vts.columns) 

            

        

    def IV(self,thres=0.02):

        e=eda(self.train,self.targetVarName,self.event)

        iv_values=e.metrics()

        print (iv_values)

        for k,v in iv_values.copy().items():

            if(v<thres):

                del iv_values[k]           

        return list(iv_values.keys())                             

    

        

    def bininfo(self,method='IV',flag=1):

        num = [x for x in self.train.columns if self.train[x].dtypes != 'object']

        cat = [x for x in self.train.columns if self.train[x].dtypes == 'object']

        #cat=df6.select_dtypes(exclude=["number","bool_","object_"])

        numdata = self.train[num]

        catdata = self.train[cat]

        utils = importr('utils')

        while True:

            try:

                smbinning = importr('smbinning')

                woeBinning = importr('woeBinning')

                InformationValue = importr('InformationValue')

                foreach = importr('foreach')

                doParallel = importr('doParallel')

            except Exception as e:

                                 

                y = str(e)

                z = re.findall(r"'(.*?)'", y, re.DOTALL)[0]

                utils.install_packages(z)

            else:

                break

        while True:

            try:

                with open('woe.r', 'r') as f:

                    string = f.read()

                    woe = STAP(string, "woe")

                    print('woe module loaded')

            except FileNotFoundError:

                path = input("set directory to path: ")

                from os import chdir

                chdir(path)

                with open('woe.r', 'r') as f:

                    string = f.read()

                    woe = STAP(string, "woe")

                    print('woe module loaded')

            else:

                break

        self.IVtable = {}

        for cal in catdata.columns:

            if cal != self.targetVarName:

                pandas2ri.activate()

                nmi = pandas2ri.ri2py(woe.wdtable(catdata[cal],catdata[self.targetVarName],self.event))

                pandas2ri.deactivate()

                self.IVtable[cal]=nmi
        try:

            numdat = numdata.copy()
    
            numdat[self.targetVarName]= self.target
    
            pandas2ri.activate()
    
            entdata = pandas2ri.ri2py(woe.woebin(numdat,self.targetVarName,self.event))
    
            pandas2ri.deactivate()

            while True:
    
                try:
    
                    for cal in entdata.columns:
    
                        pandas2ri.activate()
    
                        nmi = pandas2ri.ri2py(woe.wdtable(entdata[cal],catdata[self.targetVarName],self.event))
    
                        pandas2ri.deactivate()
    
                        self.IVtable[cal]=nmi
    
                    #pandas2ri.deactivate()
    
     
    
     
    
                    print('trying Woebin')
    
                except Exception as e:
    
                    y = str(e)
    
                    z = re.findall(r"'(.*?)'", y, re.DOTALL)[0]
    
                    utils.install_packages(z)
    
                else:
    
                    break
                
        except Exception as e:
            print(e)
                

        d=self.IVtable

        if(flag==1):

            try:

                k=0

    #            engine = 'xlsxwriter'

                book = load_workbook('univariates_bivariates.xlsx')

                writer = pd.ExcelWriter('univariates_bivariates.xlsx', engine='openpyxl')

                writer.book = book

                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                for name in d.keys():

                  b=d[name]

                  b['keys']=[name]*b.shape[0]
                  sum1=b['TOTAL'].sum()
                  l1=[]
                  for i in b['TOTAL']:
                      s=i/sum1
                      l1.append(s)
                  b['PERC']=l1
                  for s1 in b['keys']:
                      s1=s1
                      break
                  b=b[['keys','CAT','TOTAL','PERC','GOODS','BADS','PCT_G','PCT_B','WOE','IV']]
                  s1=b['keys'][0]
                  b.set_index('keys',inplace=True)
#                  b=b.sort_values('PERC')
                  print('s1 is')
                  print(s1)
                  if s1 in num:
                      b['CAT']=b['CAT'].apply(lambda x:x[2:])
                  b.to_excel(writer,sheet_name='bivariates',startrow=k,startcol=0,index=True)

                  k=k+b.shape[0]+2

                writer.close()

            except:

                k=0

    #            engine = 'xlsxwriter'

                book = load_workbook('univariates_bivariates.xlsx')

                writer = pd.ExcelWriter('univariates_bivariates.xlsx', engine='openpyxl')

                writer.book = book

                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                for name in d.keys():

                  b=d[name]

                  b['keys']=[name]*b.shape[0]

                  sum1=b['TOTAL'].sum()
                  l1=[]
                  for i in b['TOTAL']:
                      s=i/sum1
                      l1.append(s)
                  b['PERC']=l1
                  for s1 in b['keys']:
                      s1=s1
                      break
#                  b.set_index('keys',inplace=True)
                  b=b[['keys','CAT','TOTAL','PERC','GOODS','BADS','PCT_G','PCT_B','WOE','IV']]
                  b.set_index('keys',inplace=True)
#                  b=b.sort_values('PERC')
                  print('s1 is')
                  print(s1)
                  if s1 in num:
                      b['CAT']=b['CAT'].apply(lambda x:x[2:])
                  b.to_excel(writer,sheet_name='bivariates',startrow=k,startcol=0,index=True)
                  k=k+b.shape[0]+2

                writer.close()

#        a.to_csv('bin.csv')

        return d

   

    

    def calculate_vif(self,thresh,X):

        datecol=[x for x in self.train.columns if self.train[x].dtypes=='datetime64[ns]']

#       

        cat=[x for x in self.train.columns if self.train[x].dtypes == 'object' and x not in datecol]

#       

        X1=[x for x in self.train.columns if self.train[x].dtypes != 'object' and x not in datecol and x not in self.targetVarName]    

#

        X=[x for x in X1 if x not in cat]

 

        X=self.train[X]

 

        dropped=True

 

        while dropped:

 

            variables = X.columns

 

            dropped = False

 

            vif = [variance_inflation_factor(X[variables].values, X.columns.get_loc(var)) for var in X.columns]
            max_vif = max(vif)

 

            if max_vif > thresh:

 

                maxloc = vif.index(max_vif)

 

                print(f'Dropping {X.columns[maxloc]} with vif=')

 

                print(max_vif)

 

                X = X.drop([X.columns.tolist()[maxloc]], axis=1)

 

                dropped=True

 

        print(X)

 

        return X

   